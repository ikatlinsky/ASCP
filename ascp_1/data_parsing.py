# coding=utf-8
"""
Класс содержит методы извлечения, обработки, структуризации данных из xlsx файла для случая одного или нескольких
источников численности популяции.
"""

import openpyxl as xl
import numpy as np

__author__ = 'ikatlinsky'

# Общие методы


def get_ln_value(value, use_ln_value):
    """
    Возвращает исходное значение если :param use_ln_value=False, иначе верен натуральный логарифм от значения.
    :param value: значение для вычисления логарифма
    :param use_ln_value: True - если надо получить логарфм от значения
    :return: результат
    """
    return np.log(value) if use_ln_value else value

# Обработка данных из одного источника


def get_single_source_wb():
    """
    Извлекает данные из 1ой вкладки xlsx файла - данные о популяции из одного источника.
    :return: объект Worksheet
    """
    wb = xl.load_workbook(filename="ASCP01.xlsx", read_only=True, use_iterators=True)
    return wb.get_sheet_by_name("Population")


def get_single_source_data(use_time_ln=False, use_population_ln=False):
    """
    Обрабаотывает данные из одного источника, группируя их по парам [год, численность], при необходимости считает
     логарфмы от данных.
    :param use_time_ln: если True - берем логарифм для годов
    :param use_population_ln: если True - берем логарифм от численности населения
    :return: numpy массив вида [[year1, population1],[year2, population2], ...]
    """
    data_pairs = []

    for row in get_single_source_wb().iter_rows('A3:B102'):
        data_pairs.append([get_ln_value(row[0].value, use_time_ln), get_ln_value(row[1].value, use_population_ln)])

    return np.asanyarray(data_pairs)

# Обработка данных из многих источников


def get_multiple_sources_wb():
    """
    Извлекает данные из 2ой вкладки xlsx файла - данные о популяции из многих источников.
    :return: объект Worksheet
    """
    wb = xl.load_workbook(filename="ASCP01.xlsx", read_only=True, use_iterators=True)
    return wb.get_sheet_by_name("Different sources")


def get_multiple_sources_data(use_time_ln=False, use_population_ln=False):
    """
    Обрабаотывает данные из многих источников, группируя их по парам [год, численность], при необходимости считает
     логарфмы от данных. Кроме того, считает максимальные и инимальные значения популяции для каждого года.
    :param use_time_ln: если True - берем логарифм для годов
    :param use_population_ln: если True - берем логарифм от численности населения
    :return: массив вида [[[year1, population1], ...], [[year1, max_population1], ...], [[year1, min_population1], ...]]
    """
    data_pairs = []
    max_pairs = []
    min_pairs = []

    for row in get_multiple_sources_wb().iter_rows('A3:L83'):
        # Высичлем значения года при начале обраотки строки
        year = get_ln_value(row[0].value, use_time_ln)
        # Заводим массив для хранения знаений популяции для данного года
        values_row = []

        for num in range(1, len(row)):
            # проверяем, что значение не пустое в ячейке
            if row[num].value is not None:
                # проверяем, содержит ли значение символы '--', если да, в ячейке 2 значения, иначе дного значение
                if "--" not in str(row[num].value):
                    val = get_ln_value(row[num].value, use_population_ln)
                    data_pairs.append([year, val])
                    values_row.append(val)
                # обрабатываем 2 значения в одной ячейке
                else:
                    values = [x.strip() for x in row[num].value.split('--')]
                    data_pairs.append([year, get_ln_value(values[0], use_population_ln)])
                    data_pairs.append([year, get_ln_value(values[1], use_population_ln)])
                    values_row.append(get_ln_value(values[0], use_population_ln))
                    values_row.append(get_ln_value(values[1], use_population_ln))

        # если значения для года есть, то изем максимальное и минимальное значения для года и добавляем их в списки
        if len(values_row) > 0:
            max_pop = np.amax(values_row)
            min_pop = np.amin(values_row)

            max_pairs.append([year, max_pop])
            min_pairs.append([year, min_pop])

    return [np.asanyarray(data_pairs), np.asanyarray(max_pairs), np.asanyarray(min_pairs)]


print "Пример обработанных данных из одного источника: %s" % get_single_source_data()
print "Пример обработанных данных из разных источников: %s" % get_multiple_sources_data()